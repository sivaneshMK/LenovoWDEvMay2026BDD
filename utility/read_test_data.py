from openpyxl.reader.excel import load_workbook


class ExcelReader:
    path = "C:\\Users\\sivan\\PycharmProjects\\Lenovo_BDD_MayWD2026\\test_data\\test_data.xlsx"
    @staticmethod
    def get_test_data(column_name, row=2):

        wb = load_workbook(ExcelReader.path)
        sheet = wb.active
        header = {}

        for col in range(1,sheet.max_column+1):

            header[sheet.cell(1, col).value] = col

        return sheet.cell(row, header[column_name]).value
